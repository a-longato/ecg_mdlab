import csv
import h5py
import zipfile
import numpy as np
from tqdm import tqdm
from pathlib import Path
from openpyxl import load_workbook
from dataset_check import check_dataset
from dataset_prep import preprocess_dataset
from sklearn.preprocessing import LabelEncoder


def unzip_file(zip_filepath, output_dir):
    with zipfile.ZipFile(zip_filepath, "r") as zip_ref:
        files = zip_ref.namelist()

        for file in tqdm(files, desc="Unzipping files", unit="file"):
            zip_ref.extract(file, output_dir)
    print(f"Extracted all files to {output_dir}")


rhythm_mapping = {
    "AFIB": "AFIB",
    "AF": "AFIB",
    "SVT": "GSVT",
    "AT": "GSVT",
    "SAAWR": "GSVT",
    "ST": "GSVT",
    "AVNRT": "GSVT",
    "AVRT": "GSVT",
    "SB": "SB",
    "SR": "SR",
    "SA": "SR",
}

label_encoder = LabelEncoder()
encoded_classes = list(rhythm_mapping.values())
label_encoder.fit(encoded_classes)

dataset_dir = Path("Dataset")
dataset_dir.mkdir(exist_ok=True)

for i in range(len(label_encoder.classes_)):
    (dataset_dir / f"y_{i}").mkdir(exist_ok=True)


def read_xlsx(file_path):
    patient_dict = {}
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    header = {col: idx for idx, col in enumerate(next(ws.iter_rows(values_only=True)))}

    for row in tqdm(
        ws.iter_rows(values_only=True), desc="Reading Excel data", unit="row"
    ):
        if row[header["FileName"]] and row[header["Rhythm"]]:
            patient_id = row[header["FileName"]]
            rhythm = row[header["Rhythm"]]

            if rhythm in rhythm_mapping:
                mapped_rhythm = rhythm_mapping[rhythm]
                target = label_encoder.transform([mapped_rhythm])[0]
                patient_dict[patient_id] = target
                print(f"Total patient entries loaded: {len(patient_dict)}")

    wb.close()
    return patient_dict


def save_patient_hdf5(patient_id, sequences, target_class):
    file_path = dataset_dir / f"y_{target_class}" / f"{patient_id}.h5"
    print(f"Saving: {file_path} | Shape: {sequences.shape}")

    with h5py.File(file_path, "w") as f:
        f.create_dataset("sequences", data=sequences)


def read_csv(directory, patient_dict):
    for file in tqdm(
        Path(directory).glob("*.csv"), desc="Processing CSV files", unit="file"
    ):
        filename = file.stem

        if filename in patient_dict:
            target_class = patient_dict[filename]
            sequences = []

            with open(file, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)

                for row in reader:
                    try:
                        sequence = list(map(float, row))
                        if len(sequence) == 12:
                            sequences.append(sequence)
                    except ValueError:
                        continue

            if len(sequences) == 5000:
                sequences = np.array(sequences, dtype=np.float32)
                if not np.isnan(sequences).any():
                    save_patient_hdf5(filename, sequences, target_class)


xlsx_path = "Data/Diagnostics.xlsx"
csv_path = "Data/ECGDataDenoised/"
zip_filepath = "Data/ECGDataDenoised.zip"
output_dir = "Data/"

unzip_file(zip_filepath, output_dir)

patient_dict = read_xlsx(xlsx_path)

read_csv(csv_path, patient_dict)

check_dataset("Dataset")

preprocess_dataset("Dataset", "DataProcessed")
